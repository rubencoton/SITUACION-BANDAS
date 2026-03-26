from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import json
import logging
import os
import re
import unicodedata
from typing import Iterable

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from google.auth.transport.requests import Request as GoogleRequest
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials as UserCredentials

from src.config import Settings
from src.data_model import (
    BandRecord,
    NormalizedDataset,
    PHASE_ORDER,
    PhaseStateRecord,
    normalize_phase,
    normalize_status,
    status_to_score,
)


PUBLIC_EXPORT_TEMPLATE = (
    "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
)
DRIVE_EXPORT_TEMPLATE = (
    "https://www.googleapis.com/drive/v3/files/{sheet_id}/export"
    "?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
DRIVE_READONLY_SCOPE = "https://www.googleapis.com/auth/drive.readonly"


@dataclass(slots=True)
class WorkbookFetchResult:
    workbook_path: Path
    mode: str


class SheetsReader:
    def __init__(self, settings: Settings, logger: logging.Logger):
        self.settings = settings
        self.logger = logger

    def extract_dataset(self, sheet_id: str | None = None) -> NormalizedDataset:
        target_sheet_id = sheet_id or self.settings.google_sheet_id
        fetched = self.fetch_workbook_xlsx(target_sheet_id)
        extraction_ts = datetime.now()

        workbook = load_workbook(fetched.workbook_path, data_only=True)
        operational_sheets = [
            ws for ws in workbook.worksheets if self._is_operational_sheet(ws.title)
        ]

        bands: list[BandRecord] = []
        phases: list[PhaseStateRecord] = []

        for worksheet in operational_sheets:
            band_row, phase_rows = self._parse_sheet(worksheet, extraction_ts)
            bands.append(band_row)
            phases.extend(phase_rows)

        bands_df = pd.DataFrame([row.__dict__ for row in bands])
        phases_df = pd.DataFrame([row.__dict__ for row in phases])

        if not bands_df.empty and not phases_df.empty:
            score_avg = (
                phases_df.groupby("banda", as_index=False)["score"]
                .mean(numeric_only=True)
                .rename(columns={"score": "score_global"})
            )
            bands_df = bands_df.merge(score_avg, how="left", on="banda")
        else:
            bands_df["score_global"] = pd.Series(dtype=float)

        return NormalizedDataset(
            bands_df=bands_df,
            phases_df=phases_df,
            extraction_ts=extraction_ts,
            sheet_id=target_sheet_id,
            source_mode=fetched.mode,
        )

    def fetch_workbook_xlsx(self, sheet_id: str) -> WorkbookFetchResult:
        output_dir = self.settings.output_dir / "_downloads"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{sheet_id}.xlsx"

        if self.settings.use_public_sheet_export:
            content = self._download_public_export(sheet_id)
            if content:
                output_path.write_bytes(content)
                return WorkbookFetchResult(output_path, mode="public_export")

        if self.settings.has_google_credentials:
            content = self._download_drive_export(sheet_id)
            if content:
                output_path.write_bytes(content)
                return WorkbookFetchResult(output_path, mode="drive_api_export")

        raise RuntimeError(
            "No se pudo descargar la Google Sheet. "
            "Revisa permisos o credenciales de Google."
        )

    def _download_public_export(self, sheet_id: str) -> bytes | None:
        url = PUBLIC_EXPORT_TEMPLATE.format(sheet_id=sheet_id)
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            data = response.content
            if data.startswith(b"PK"):
                self.logger.info(
                    "Lectura de Google Sheet por export publico XLSX completada."
                )
                return data
            self.logger.warning("Export publico no devolvio XLSX valido.")
            return None
        except requests.RequestException as exc:
            self.logger.warning("Fallo export publico: %s", exc)
            return None

    def _download_drive_export(self, sheet_id: str) -> bytes | None:
        token = self._build_google_bearer_token()
        if not token:
            return None
        headers = {"Authorization": f"Bearer {token}"}
        try:
            response = requests.get(
                DRIVE_EXPORT_TEMPLATE.format(sheet_id=sheet_id),
                headers=headers,
                timeout=30,
            )
            response.raise_for_status()
            data = response.content
            if data.startswith(b"PK"):
                self.logger.info("Lectura de Google Sheet por Drive API completada.")
                return data
            self.logger.warning("Drive export no devolvio XLSX valido.")
            return None
        except requests.RequestException as exc:
            self.logger.warning("Fallo Drive API export: %s", exc)
            return None

    def _build_google_bearer_token(self) -> str | None:
        if self.settings.google_service_account_file:
            path = Path(self.settings.google_service_account_file)
            if not path.exists():
                self.logger.warning("No existe service account file: %s", path)
                return None
            creds = service_account.Credentials.from_service_account_file(
                str(path), scopes=[DRIVE_READONLY_SCOPE]
            )
            creds.refresh(GoogleRequest())
            return creds.token

        if self.settings.google_oauth_user_file:
            path = Path(self.settings.google_oauth_user_file)
            if not path.exists():
                self.logger.warning("No existe OAuth user file: %s", path)
                return None
            creds = UserCredentials.from_authorized_user_file(
                str(path), scopes=[DRIVE_READONLY_SCOPE]
            )
            creds.refresh(GoogleRequest())
            return creds.token

        if self.settings.google_clasp_profile:
            clasp_path = (
                Path(self.settings.google_clasp_file).expanduser()
                if self.settings.google_clasp_file
                else Path.home() / ".clasprc.json"
            )
            if not clasp_path.exists():
                self.logger.warning("No existe archivo .clasprc: %s", clasp_path)
                return None
            content = json.loads(clasp_path.read_text(encoding="utf-8"))
            token_obj = (content.get("tokens") or {}).get(
                self.settings.google_clasp_profile
            )
            if not token_obj:
                self.logger.warning(
                    "No existe perfil %s en %s.",
                    self.settings.google_clasp_profile,
                    clasp_path,
                )
                return None
            try:
                response = requests.post(
                    "https://oauth2.googleapis.com/token",
                    data={
                        "client_id": token_obj["client_id"],
                        "client_secret": token_obj["client_secret"],
                        "refresh_token": token_obj["refresh_token"],
                        "grant_type": "refresh_token",
                    },
                    timeout=30,
                )
                response.raise_for_status()
                return response.json().get("access_token")
            except requests.RequestException as exc:
                self.logger.warning("No se pudo refrescar token .clasprc: %s", exc)
                return None
        return None

    @staticmethod
    def _is_operational_sheet(sheet_name: str) -> bool:
        normalized = re.sub(r"\s+", "", sheet_name.upper())
        if normalized in {"MODELO", "PLANTILLA", "TEMPLATE", "_MODELO"}:
            return False
        if sheet_name.startswith("_"):
            return False
        return True

    def _parse_sheet(self, ws, extraction_ts: datetime) -> tuple[BandRecord, list[PhaseStateRecord]]:
        banda = self._read_field(
            ws,
            primary_coords=["B1", "C1"],
            label_aliases=["NOMBRE BANDA", "BANDA"],
        )
        correo = self._read_field(
            ws,
            primary_coords=["B2", "C2"],
            label_aliases=["CORREO", "EMAIL"],
        )
        asunto = self._read_field(
            ws,
            primary_coords=["B3", "C3"],
            label_aliases=["ASUNTO"],
        )
        mensaje = self._read_field(
            ws,
            primary_coords=["B4", "C4"],
            label_aliases=["MENSAJE PERSONALIZADO", "MENSAJE"],
        )
        last_sent_raw = ws["F2"].value
        if last_sent_raw in (None, ""):
            label_cell = self._find_label_cell(
                ws, ["ULTIMO ENVIO:", "ULTIMO ENVIO", "ULTIMO ENVÍO", "ULTIMO ENVÍO:"]
            )
            if label_cell:
                row, col = label_cell
                for step in range(1, 4):
                    candidate = ws.cell(row=row, column=col + step).value
                    if candidate not in (None, ""):
                        last_sent_raw = candidate
                        break
        ultimo_envio = self._parse_datetime_value(last_sent_raw)
        observaciones = self._read_observaciones(ws)

        band_record = BandRecord(
            banda=banda or ws.title.strip(),
            correo_principal=correo,
            asunto=asunto,
            mensaje_personalizado=mensaje,
            ultimo_envio=ultimo_envio,
            observaciones=observaciones,
            nombre_hoja=ws.title,
            fecha_extraccion=extraction_ts.isoformat(),
        )

        states = self._read_phase_states(ws, band_record.banda, extraction_ts)
        return band_record, states

    @staticmethod
    def _clean(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return re.sub(r"\s+", " ", text)

    def _read_field(
        self,
        ws,
        primary_coords: Iterable[str],
        label_aliases: Iterable[str],
    ) -> str:
        for coord in primary_coords:
            val = self._clean(ws[coord].value)
            if val:
                return val

        label_cell = self._find_label_cell(ws, label_aliases)
        if label_cell is None:
            return ""

        row, col = label_cell
        for step in range(1, 5):
            value = self._clean(ws.cell(row=row, column=col + step).value)
            if value:
                return value
        return ""

    def _find_label_cell(self, ws, aliases: Iterable[str]) -> tuple[int, int] | None:
        normalized_aliases = {self._normalize_label_text(alias) for alias in aliases}
        for row in range(1, 35):
            for col in range(1, 10):
                val = self._normalize_label_text(
                    self._clean(ws.cell(row=row, column=col).value)
                )
                if not val:
                    continue
                if val in normalized_aliases:
                    return row, col
        return None

    @staticmethod
    def _normalize_label_text(value: str) -> str:
        if not value:
            return ""
        no_accents = "".join(
            c
            for c in unicodedata.normalize("NFD", value)
            if unicodedata.category(c) != "Mn"
        )
        no_accents = no_accents.upper().strip()
        no_accents = no_accents.replace(":", "")
        no_accents = re.sub(r"\s+", " ", no_accents)
        return no_accents

    def _read_observaciones(self, ws) -> str:
        values: list[str] = []
        for row in range(14, 18):
            for col in range(1, 7):
                cell = self._clean(ws.cell(row=row, column=col).value)
                if cell and cell not in {"OBSERVACIONES"}:
                    values.append(cell)
        if values:
            # Mantener orden y eliminar duplicados inmediatos de rangos fusionados.
            dedup: list[str] = []
            for item in values:
                if not dedup or dedup[-1] != item:
                    dedup.append(item)
            return " | ".join(dedup)

        label_cell = self._find_label_cell(ws, ["OBSERVACIONES"])
        if not label_cell:
            return ""
        row, col = label_cell
        for offset in range(1, 6):
            candidate = self._clean(ws.cell(row=row + offset, column=col).value)
            if candidate:
                return candidate
        return ""

    def _read_phase_states(
        self, ws, banda: str, extraction_ts: datetime
    ) -> list[PhaseStateRecord]:
        phase_status_map: dict[str, str] = {}

        for row in range(7, 12):
            phase_label = normalize_phase(self._clean(ws.cell(row=row, column=1).value))
            status_label = normalize_status(self._clean(ws.cell(row=row, column=2).value))
            if phase_label == "DESCONOCIDA":
                continue
            if phase_label in PHASE_ORDER:
                phase_status_map[phase_label] = status_label

        # Fallback semantico si la lectura por coordenadas no fue suficiente.
        if len(phase_status_map) < 3:
            for row in range(1, 40):
                for col in range(1, 8):
                    raw_phase = self._clean(ws.cell(row=row, column=col).value)
                    phase_label = normalize_phase(raw_phase)
                    if phase_label not in PHASE_ORDER:
                        continue
                    for candidate_col in range(col + 1, col + 4):
                        raw_status = self._clean(
                            ws.cell(row=row, column=candidate_col).value
                        )
                        status_label = normalize_status(raw_status)
                        if status_label != "DESCONOCIDO":
                            phase_status_map[phase_label] = status_label
                            break

        rows: list[PhaseStateRecord] = []
        for phase in PHASE_ORDER:
            status = phase_status_map.get(phase, "DESCONOCIDO")
            rows.append(
                PhaseStateRecord(
                    banda=banda,
                    fase=phase,
                    estado=status,
                    score=status_to_score(status),
                    fecha_extraccion=extraction_ts.isoformat(),
                    nombre_hoja=ws.title,
                )
            )
        return rows

    def _parse_datetime_value(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, (int, float)):
            try:
                dt = from_excel(value)
                return dt.isoformat()
            except Exception:
                return str(value)
        parsed = pd.to_datetime(str(value), errors="coerce", dayfirst=True)
        if pd.isna(parsed):
            return self._clean(value)
        return parsed.to_pydatetime().isoformat()
